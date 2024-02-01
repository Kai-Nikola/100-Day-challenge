import openpyxl
from openpyxl.chart import BarChart, Reference

wb = openpyxl.load_workbook("marks_data.xlsx")

wb.active
sheet = wb["Sheet1"]

row_len = sheet.max_row
column_len = 3


print("Enter marks subject wise to print marks in bar chart")

hindi = int(input("Hindi "))
marathi = int(input("Marathi "))
english = int(input("English "))
mathematics = int(input("Mathematics "))
science = int(input("Science "))
ss = int(input("Social Science "))

sheet.cell(row=1, column=1).value = "Subject"
sheet.cell(row=1, column=2).value = "Marks Obtained"
sheet.cell(row=1, column=3).value = "Total Marks"

# Adding Subjects
sheet.cell(row=2, column=1).value = "Hindi"
sheet.cell(row=3, column=1).value = "Marathi"
sheet.cell(row=4, column=1).value = "english".title()
sheet.cell(row=5, column=1).value = "mathematics".title()
sheet.cell(row=6,column=1).value = "science".title()
sheet.cell(row=7, column=1).value = "Social Science"

# Adding Marks
sheet.cell(row=2, column=2).value = hindi
sheet.cell(row=3, column=2).value = marathi
sheet.cell(row=4, column=2).value = english
sheet.cell(row=5, column=2).value = mathematics
sheet.cell(row=6,column=2).value = science
sheet.cell(row=7, column=2).value = ss

# Total Marks

for mark in range(2, 8):
    sheet.cell(row=mark, column=3).value = 100

selection_values = Reference(sheet, min_row=2, max_row=row_len, min_col=2, max_col=column_len)

chart = BarChart()
chart.add_data(selection_values)
sheet.add_chart(chart, 'e2')

wb.save("marks_data.xlsx")
wb.close()